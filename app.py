# importando as bibliotecas do projeto
import customtkinter as ctk 
from tkinter import *
from tkinter import messagebox
import openpyxl
import pathlib
from openpyxl import Workbook

# criando a interface do sistema
ctk.set_appearance_mode("Light")
ctk.set_default_color_theme("blue")

class App(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.layout_config()
        self.todo_sistema()
        
    # definindo configurações de layout
    def layout_config(self):
        self.title("Controle Estoque - OCM")
        self.geometry("700x500")

    # definindo elementos da interface
    def todo_sistema(self):
        frame = ctk.CTkFrame(self, width=700, height=50, corner_radius=0, bg_color="teal", fg_color="teal").place(x=0, y=10)
        title = ctk.CTkLabel(frame, text="Controle de Estoque - OCM", font=("Century Gothic bold", 22), text_color="#fff")
        span = ctk.CTkLabel(self, text="Preencha todos os campos!", font=("Century Gothic bold", 14), text_color=["#000", "#fff"]).place(x=190, y=10)

        ficheiro = pathlib.Path("banco.xlsx")

        if not ficheiro.exists():
            workbook = Workbook()
            folha = workbook.active
            folha['A1'] = "Nome Completo"
            folha['B1'] = "Contato"
            folha['C1'] = "Idade"  
            folha['D1'] = "Gênero"
            folha['E1'] = "Endereço"
            folha['F1'] = "Observações"

            workbook.save("banco.xlsx")

        def submit():
            name = name_value.get()
            contact = contact_value.get()
            age = age_value.get()
            gender = gender_combobox.get()
            address = address_value.get()
            obs = obs_entry.get(1.0, END)

            workbook = openpyxl.load_workbook('banco.xlsx')
            folha = workbook.active
            folha.cell(column=1, row=folha.max_row+1, value=name)
            folha.cell(column=2, row=folha.max_row, value=contact)
            folha.cell(column=3, row=folha.max_row, value=age)
            folha.cell(column=4, row=folha.max_row, value=gender)
            folha.cell(column=5, row=folha.max_row, value=address)
            folha.cell(column=6, row=folha.max_row, value=obs)

            workbook.save("banco.xlsx")
            messagebox.showinfo("Sistema", "Dados Salvos com Sucesso!")

        def clear():
            name_value.set("")
            contact_value.set("")
            age_value.set("")
            address_value.set("")
            obs_entry.delete(1.0, END)

        name_value = StringVar()
        contact_value = StringVar()  
        age_value = StringVar()
        address_value = StringVar()

        # entrys
        name_entry = ctk.CTkEntry(self, width=350, textvariable=name_value, font=("Century Gothic bold", 16), fg_color="transparent")
        contact_entry = ctk.CTkEntry(self, width=200, textvariable=contact_value, font=("Century Gothic bold", 16), fg_color="transparent")
        age_entry = ctk.CTkEntry(self, width=150, textvariable=age_value, font=("Century Gothic bold", 16), fg_color="transparent")
        address_entry = ctk.CTkEntry(self, width=200, textvariable=address_value, font=("Century Gothic bold", 16), fg_color="transparent")

        # combobox
        gender_combobox = ctk.CTkComboBox(self, values=["Feminino", "Masculino", "Infantil"], font=("Century Gothic bold", 14))
        gender_combobox.set("Feminino")

        # entrada de observações
        obs_entry = ctk.CTkTextbox(self, width=500, height=150, font=("Arial", 18), border_color="#aaa", border_width=2, fg_color="transparent")

        # Campos de informações
        lb_name = ctk.CTkLabel(self, text="Nome Completo:", font=("Century Gothic bold", 14), text_color=["#000", "#fff"])
        lb_contact = ctk.CTkLabel(self, text="Contato:", font=("Century Gothic bold", 14), text_color=["#000", "#fff"])
        lb_age = ctk.CTkLabel(self, text="Idade:", font=("Century Gothic bold", 14), text_color=["#000", "#fff"])
        lb_gender = ctk.CTkLabel(self, text="Género:", font=("Century Gothic bold", 14), text_color=["#000", "#fff"])
        lb_address = ctk.CTkLabel(self, text="Endereço:", font=("Century Gothic bold", 14), text_color=["#000", "#fff"])
        lb_obs = ctk.CTkLabel(self, text="Observações", font=("Century Gothic bold", 14), text_color=["#000", "#fff"])

        # botao banco de dados
        create_button = ctk.CTkButton(self, text="Abrir Banco de Dados".upper(), font=("Century Gothic bold", 14), bg_color="transparent", fg_color="#8B4513", hover_color="#B8860B", text_color="#fff", command=self.show_window)
        create_button.place(x=50, y=420)

        # botoes do sistema
        btn_submit = ctk.CTkButton(self, text="Salvar Dados".upper(), command=submit, fg_color="#008000", hover_color="#2E8B57").place(x=300, y=420)
        btn_submit = ctk.CTkButton(self, text="Limpar Campos".upper(), command=clear, fg_color="#FF0000", hover_color="#FF7F50").place(x=500, y=420)

        # posicionamento dos elementos
        lb_name.place(x=50, y=120)
        name_entry.place(x=50, y=150)

        lb_contact.place(x=450, y=120)
        contact_entry.place(x=450, y=150)

        lb_age.place(x=300, y=190)
        age_entry.place(x=300, y=220)

        lb_gender.place(x=510, y=190)
        gender_combobox.place(x=510, y=220)

        lb_address.place(x=50, y=190)
        address_entry.place(x=50, y=220)

        lb_obs.place(x=50, y=260)
        obs_entry.place(x=150, y=260)

        
    def change_apm(self, nova_aparencia):
        ctk.set_appearance_mode(nova_aparencia)

    def show_window(self):
        new_window = Toplevel(self)
        new_window.title("Banco de dados")
        new_window.geometry("300x200")
        new_label = ctk.CTkLabel(new_window, text="Banco de dados")
        new_label.pack()

# janela principal - sempre no fim do codigo
if __name__ == "__main__":
    app = App()
    app.mainloop()
